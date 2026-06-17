"""Autorrelleno de la taxonomía del catálogo en Supabase desde el Excel BDD + Morningstar.

Campos (fund_groups): tipo_activo, geografia (regla developed/acwi en RV), categoria_rf
(tipo de bono, de Morningstar), plazo (CP/MP/LP del Excel), srri (Excel), estilo
(Morningstar), caracteristicas_especiales (multi, derivadas). Excel = fuente de tus
campos; Morningstar = enriquecidos.

Uso: python -m tools.sync_taxonomy [--apply]
"""
from __future__ import annotations
import re, sys
from pathlib import Path
from collections import Counter

EXCEL = r"C:\Users\RafaelGarcía\OneDrive - Nazca\Rafa\Personal\Asesoría Financiera\Operativa\BDD\20260513_Listado Fondos Def (version valores).xlsm"
_ISIN = re.compile(r"^[A-Z]{2}[A-Z0-9]{10}$")


def read_excel():
    import openpyxl
    wb = openpyxl.load_workbook(EXCEL, data_only=True, read_only=True)
    ws = wb["Listado fondos"]
    out = {}
    for r in range(4, ws.max_row + 1):
        i = str(ws.cell(r, 17).value or ws.cell(r, 2).value or "").strip().upper()
        if not _ISIN.match(i):
            continue
        out[i] = {"tipo": ws.cell(r, 5).value, "geo": ws.cell(r, 6).value,
                  "cat_gestion": ws.cell(r, 4).value, "srri": ws.cell(r, 59).value,
                  "nombre": ws.cell(r, 18).value}
    return out


def m_tipo(tipo, cg):
    t = (tipo or "").strip(); c = (cg or "").strip()
    if c in ("Monetario", "Ultra Short"):
        return "Monetario"
    if t.startswith("RV"): return "RV"
    if t.startswith("RF"): return "RF"
    if t.startswith("Mixto"): return "Mixtos"
    if t.startswith("Alternativ"): return "Alternativos"
    if "materias" in t.lower(): return "Materias_primas"
    return None


def m_plazo(tipo):
    t = (tipo or "")
    if re.search(r"\bCP\b", t): return "Corto"
    if re.search(r"\bMP\b", t): return "Medio"
    if re.search(r"\bLP\b", t): return "Largo"
    return None


def m_geo(geo, tipo_act, nombre):
    g = (geo or "").strip()
    if tipo_act == "RV":
        n = (nombre or "").lower()
        if g.lower() == "developed" or "developed world" in n or "msci world" in n or \
           (re.search(r"\bworld\b", n) and "all" not in n and "acwi" not in n):
            return "Developed"
        if g.lower() == "global" or "acwi" in n or "all country" in n or "all-world" in n:
            return "ACWI"
        return g or None
    return g or None


def m_cat_rf(cm, tipo_act):
    if tipo_act != "RF":
        return None
    cm = (cm or "").lower()
    if "government" in cm or "govt" in cm or "inflation" in cm:
        return "Gubernamental"
    if "corporate" in cm or "subordinated" in cm or "high yield" in cm:
        return "Corporativa"
    if "flexible" in cm or "diversified" in cm or "other bond" in cm:
        return "Flexible/Diversificada"
    return None


def m_tags(tipo, cg, cm, nombre):
    t = (tipo or ""); c = (cg or ""); m = (cm or "").lower(); n = (nombre or "").lower()
    tags = set()
    if re.search(r"\bHY\b", t) or "high yield" in m: tags.add("High Yield")
    if "catastrofe" in t.lower() or "insurance-linked" in n or " ils" in n: tags.add("ILS/Catástrofe")
    if "retorno absoluto" in t.lower(): tags.add("Retorno absoluto")
    if c in ("Indexado", "ETF"): tags.add("Indexado/ETF")
    if c == "Sectoriales": tags.add("Sectorial")
    if "small" in m or "mid-cap" in m or "small/mid" in m: tags.add("Small/Mid Caps")
    if "property" in m: tags.add("REITs")
    if "hedged" in m or "eurhdg" in n or "eur hdg" in n: tags.add("Cubre divisa")
    return sorted(tags)


def m_srri(v):
    try:
        n = int(float(v)); return n if 1 <= n <= 7 else None
    except Exception:
        return None


def main():
    apply = "--apply" in sys.argv
    from dotenv import load_dotenv
    load_dotenv(Path(__file__).resolve().parent.parent / ".env")
    from tools.supabase_client import get_client
    c = get_client()
    xl = read_excel()
    funds = c.table("funds").select("isin,fund_group_id").limit(5000).execute().data
    groups = {g["fund_group_id"]: g for g in c.table("fund_groups").select(
        "fund_group_id,categoria_morningstar,nombre_base,tipo_activo,geografia,estilo").limit(5000).execute().data}
    g2isins = {}
    for f in funds:
        g2isins.setdefault(f["fund_group_id"], []).append(f["isin"])

    dist = {k: Counter() for k in ("tipo", "geo", "cat_rf", "plazo", "srri", "tags")}
    n_upd = 0
    samples = []
    for gid, isins in g2isins.items():
        g = groups.get(gid, {})
        xrec = next((xl[i] for i in isins if i in xl), None)
        cm = g.get("categoria_morningstar")
        nombre = (xrec or {}).get("nombre") or g.get("nombre_base")
        # Excel manda; si el fondo no está en el Excel, usamos lo que ya hay en
        # Supabase como fallback (p.ej. Cobas) para igual aplicar la regla dev/acwi.
        tipo = m_tipo((xrec or {}).get("tipo"), (xrec or {}).get("cat_gestion")) or g.get("tipo_activo")
        geo = m_geo((xrec or {}).get("geo") or g.get("geografia"), tipo, nombre)
        cat_rf = m_cat_rf(cm, tipo)
        plazo = m_plazo((xrec or {}).get("tipo"))
        srri = m_srri((xrec or {}).get("srri")) if xrec else None
        tags = m_tags((xrec or {}).get("tipo"), (xrec or {}).get("cat_gestion"), cm, nombre)
        upd = {}
        if tipo: upd["tipo_activo"] = tipo; dist["tipo"][tipo] += 1
        if geo: upd["geografia"] = geo; dist["geo"][geo] += 1
        if cat_rf: upd["categoria_rf"] = cat_rf; dist["cat_rf"][cat_rf] += 1
        if plazo: upd["plazo"] = plazo; dist["plazo"][plazo] += 1
        if srri: upd["srri"] = srri; dist["srri"][srri] += 1
        if tags: upd["caracteristicas_especiales"] = tags; [dist["tags"].update([t]) for t in tags]
        if upd:
            n_upd += 1
            if len(samples) < 8:
                samples.append((nombre, upd))
            if apply:
                try:
                    c.table("fund_groups").update(upd).eq("fund_group_id", gid).execute()
                except Exception as e:
                    msg = str(e)
                    key = "constraint " + (re.search(r'"(chk_\w+)"', msg).group(1) if re.search(r'"(chk_\w+)"', msg) else msg[:40])
                    dist.setdefault("_errores", Counter())[key] += 1
    print(f"{'APLICADO' if apply else 'DRY-RUN'}: {n_upd} grupos actualizados\n")
    for k, ctr in dist.items():
        print(f"  {k}: {dict(ctr.most_common())}")
    print("\n  muestras:")
    for nm, upd in samples:
        print(f"    {(nm or '')[:30]:30} {upd}")


if __name__ == "__main__":
    main()
