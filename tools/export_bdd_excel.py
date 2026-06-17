"""Exporta el catálogo de Supabase a un Excel maestro (BDD v2) — espejo 100% de Supabase.

Una fila por clase (fondo). Columnas: identidad + taxonomía (tipo/geo/catRF/plazo/
estilo/características/SRRI/Morningstar) + clasificación/opinión/encaje + cuantitativos
por clase (divisa/distribución/TER/comisión/éxito).

Uso: python -m tools.export_bdd_excel [ruta_salida.xlsx]
"""
from __future__ import annotations
import sys
from pathlib import Path
from datetime import datetime

# Ruta por defecto del Excel espejo (OneDrive BDD).
DEFAULT_OUT = (Path(r"C:\Users\RafaelGarcía\OneDrive - Nazca\Rafa\Personal"
                    r"\Asesoría Financiera\Operativa\BDD")
               / "20260617_BDD Fondos v2 (espejo Supabase).xlsx")


def regenerate(out: str | None = None, quiet: bool = False) -> str | None:
    """Regenera el Excel espejo desde Supabase. Devuelve la ruta o None si falla.

    Best-effort: NO lanza si Supabase no está o el fichero está bloqueado (Excel
    abierto) — devuelve None y, si quiet=False, imprime el motivo. Pensado para
    llamarse tras cada categorización (web_server) y al final del pipeline.
    """
    try:
        return _build(out, quiet)
    except PermissionError:
        if not quiet:
            print("[WARN] Excel espejo bloqueado (¿abierto en Excel?). No se regeneró.")
        return None
    except Exception as e:
        if not quiet:
            print(f"[WARN] No se pudo regenerar el Excel espejo: {str(e)[:160]}")
        return None


def _build(out: str | None = None, quiet: bool = False) -> str:
    from dotenv import load_dotenv
    load_dotenv(Path(__file__).resolve().parent.parent / ".env")
    from tools.supabase_client import get_client
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    c = get_client()
    funds = c.table("funds").select("*").limit(5000).execute().data
    groups = {g["fund_group_id"]: g for g in c.table("fund_groups").select("*").limit(5000).execute().data}

    cols = [
        ("ISIN", lambda f, g: f.get("isin")),
        ("Fondo", lambda f, g: f.get("nombre_clase") or g.get("nombre_base")),
        ("Gestora", lambda f, g: g.get("gestora")),
        ("Tipo activo", lambda f, g: g.get("tipo_activo")),
        ("Geografía", lambda f, g: g.get("geografia")),
        ("Categoría RF", lambda f, g: g.get("categoria_rf")),
        ("Plazo", lambda f, g: g.get("plazo")),
        ("Estilo", lambda f, g: g.get("estilo")),
        ("Características", lambda f, g: ", ".join(g.get("caracteristicas_especiales") or [])),
        ("Categoría Morningstar", lambda f, g: g.get("categoria_morningstar")),
        ("SRRI", lambda f, g: g.get("srri")),
        ("Clasificación", lambda f, g: f.get("clasificacion_user")),
        ("Opinión", lambda f, g: f.get("opinion_user")),
        ("Encaje", lambda f, g: f.get("encaje_texto")),
        ("Divisa", lambda f, g: f.get("divisa")),
        ("Distribución", lambda f, g: f.get("distribucion")),
        ("TER %", lambda f, g: f.get("ter_pct")),
        ("Comisión gestión %", lambda f, g: f.get("comision_gestion_pct")),
        ("Comisión éxito", lambda f, g: ("Sí" if (f.get("comision_exito_pct") or 0) else "No") if f.get("comision_exito_pct") is not None else None),
        ("Año inicio", lambda f, g: str(f.get("fecha_creacion_clase"))[:4] if f.get("fecha_creacion_clase") else None),
        ("AUM (M€)", lambda f, g: g.get("aum_meur")),
        ("Años", lambda f, g: g.get("años_antiguedad")),
        ("Rating MS", lambda f, g: (((g.get("portfolio_metrics_jsonb") or {}).get("morningstar") or {}).get("estrellas"))),
        ("Medalist", lambda f, g: (((g.get("portfolio_metrics_jsonb") or {}).get("morningstar") or {}).get("medalist"))),
        ("CAGR % (MS daily)", lambda f, g: (g.get("rendimiento_jsonb") or {}).get("cagr_desde_inicio")),
        ("Volatilidad %", lambda f, g: (g.get("rendimiento_jsonb") or {}).get("volatilidad")),
        ("Volat. 5A %", lambda f, g: (g.get("rendimiento_jsonb") or {}).get("volatilidad_5a")),
        ("Max DD %", lambda f, g: (g.get("rendimiento_jsonb") or {}).get("max_drawdown")),
        ("Peor año %", lambda f, g: (g.get("rendimiento_jsonb") or {}).get("peor_anio")),
        ("Mejor año %", lambda f, g: (g.get("rendimiento_jsonb") or {}).get("mejor_anio")),
        ("Rent. 3A anual %", lambda f, g: (g.get("rendimiento_jsonb") or {}).get("rentab_3a")),
        ("Rent. 5A anual %", lambda f, g: (g.get("rendimiento_jsonb") or {}).get("rentab_5a")),
        ("% RV", lambda f, g: (((g.get("portfolio_metrics_jsonb") or {}).get("myinvestor") or {}).get("asset_allocation") or {}).get("equity")),
        ("% RF", lambda f, g: (((g.get("portfolio_metrics_jsonb") or {}).get("myinvestor") or {}).get("asset_allocation") or {}).get("bond")),
        ("% Liquidez", lambda f, g: (((g.get("portfolio_metrics_jsonb") or {}).get("myinvestor") or {}).get("asset_allocation") or {}).get("cash")),
        ("Análisis", lambda f, g: "Sí" if f.get("has_qualitative_analysis") else "No"),
        ("Disp. MyInvestor", lambda f, g: "Sí" if "MyInvestor" in (f.get("broker_disponible") or []) else "No"),
        ("Brokers", lambda f, g: ", ".join(f.get("broker_disponible") or [])),
        ("fund_group_id", lambda f, g: f.get("fund_group_id")),
    ]
    # ordenar por grupo (nombre_base) y luego por clase
    funds.sort(key=lambda f: ((groups.get(f.get("fund_group_id"), {}).get("nombre_base") or "zzz").lower(), f.get("isin") or ""))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Catálogo (Supabase)"
    hdr_fill = PatternFill("solid", fgColor="003A75")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    for j, (name, _) in enumerate(cols, 1):
        cell = ws.cell(1, j, name)
        cell.fill = hdr_fill; cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for i, f in enumerate(funds, 2):
        g = groups.get(f.get("fund_group_id"), {})
        for j, (_, fn) in enumerate(cols, 1):
            ws.cell(i, j, fn(f, g))
    ws.freeze_panes = "A2"
    widths = [14, 38, 24, 12, 11, 16, 8, 10, 28, 26, 6, 13, 40, 24, 8, 12, 8, 14, 12, 9, 10, 6, 9, 18, 38]
    from openpyxl.utils import get_column_letter
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = w

    out = out or str(DEFAULT_OUT)
    wb.save(out)
    if not quiet:
        print(f"Excel v2 generado: {out}  ({len(funds)} filas)")
    return out


def main():
    out = sys.argv[1] if len(sys.argv) > 1 and not sys.argv[1].startswith("--") else None
    regenerate(out)


if __name__ == "__main__":
    main()
